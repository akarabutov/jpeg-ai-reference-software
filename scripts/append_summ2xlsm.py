from openpyxl import load_workbook
import argparse
import os

def get_args():
    cur_file_path=os.path.dirname(os.path.abspath(__file__))
    parser = argparse.ArgumentParser()
    parser.add_argument('summary_file', type=str, help="Path to summary file")
    parser.add_argument('output_file', type=str, help="Path to output file")
    parser.add_argument('--template', type=str, default=os.path.join(cur_file_path, os.pardir, "docs", "template.xlsm"), help="Path to template Excel file")
    parser.add_argument('--page', type=str, default='Test', help="Name of the target page")
    parser.add_argument('--start-row', type=int, default=3, help="Row of starting cell")
    parser.add_argument('--start-col', type=int, default=32, help="Column of starting cell")
    parser.add_argument('--anchor', type=str, default=None, help="Set anchor on summary page")
    return parser.parse_args()


def convert_to_float(d):
    try:
        ans = float(d)
    except:
        ans = d
    return ans

def append_summary(summary_file: str, output_file: str, page:str='Test', start_row:int=3, start_col:int=32, template:str=None, anchor:str=None) -> None:
    if template is None:
        cur_file_path=os.path.dirname(os.path.abspath(__file__))
        template = os.path.join(cur_file_path, os.pardir, "docs", "template.xlsm")
    wb = load_workbook(template,keep_vba=True)
    page = wb[page]
    with open(summary_file, 'r') as f:
        cur_r = start_row
        for l in f:
            cur_c = start_col
            for c in l.split('\t'):
                c = convert_to_float(c)
                page.cell(row=cur_r, column=cur_c, value=c)       
                cur_c += 1
            cur_r += 1
    if anchor is not None:
        page_summary = wb["Summary"]
        page_summary.cell(row=1, column=2, value=anchor)
    wb.save(output_file)
   
def main():
    args = get_args()
    append_summary(**vars(args))
    
    
if __name__ == "__main__":
    main()